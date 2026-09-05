"""Export service (seção 17): turns a `ReportDocument` into CSV, Excel or PDF
bytes. Every report type in `report_service.py` builds the same generic
document — one title, one subtitle line (period/filters), a header row and
data rows — so the three renderers here only need to be written once and
reused by every report, instead of every report type reinventing its own
Excel/PDF layout.
"""
import csv
import io
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_ACCENT = "4338CA"
_ACCENT_HEX = colors.HexColor(f"#{_ACCENT}")
_BORDER_HEX = colors.HexColor("#E3E6EC")
_ROW_ALT_HEX = colors.HexColor("#F1F3F6")


@dataclass
class ReportDocument:
    """A generic tabular report: a header (title/subtitle/company/generated
    at) plus columns and rows, all already formatted as display strings —
    formatting decisions (dates, percentages, currency) belong to the code
    that builds the document, not to the exporters below.
    """

    title: str
    subtitle: str
    tenant_name: str
    columns: list[str]
    rows: list[list[str]] = field(default_factory=list)
    generated_at: datetime = field(default_factory=datetime.utcnow)


def to_csv(doc: ReportDocument) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([doc.title])
    writer.writerow([doc.tenant_name])
    writer.writerow([doc.subtitle])
    writer.writerow([f"Gerado em {doc.generated_at.strftime('%d/%m/%Y %H:%M')}"])
    writer.writerow([])
    writer.writerow(doc.columns)
    writer.writerows(doc.rows)
    # BOM (utf-8-sig): sem isso o Excel abre acentos como lixo ao clicar
    # duas vezes no CSV no Windows — não é opcional para um relatório PT-BR.
    return buffer.getvalue().encode("utf-8-sig")


def to_excel(doc: ReportDocument) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"

    last_col = max(len(doc.columns), 1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = sheet.cell(1, 1, doc.title)
    title_cell.font = Font(size=15, bold=True, color=_ACCENT)

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    sheet.cell(2, 1, doc.tenant_name).font = Font(size=10, bold=True)

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    sheet.cell(3, 1, doc.subtitle).font = Font(size=10, italic=True, color="5B6474")

    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    generated_cell = sheet.cell(4, 1, f"Gerado em {doc.generated_at.strftime('%d/%m/%Y %H:%M')}")
    generated_cell.font = Font(size=9, color="98A2B3")

    header_row = 6
    for col_index, column_name in enumerate(doc.columns, start=1):
        cell = sheet.cell(header_row, col_index, column_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, row in enumerate(doc.rows):
        row_index = header_row + 1 + row_offset
        fill = PatternFill("solid", fgColor="F1F3F6") if row_offset % 2 else None
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, col_index, value)
            if fill:
                cell.fill = fill

    for col_index, column_name in enumerate(doc.columns, start=1):
        widest = max([len(column_name)] + [len(str(row[col_index - 1])) for row in doc.rows], default=10)
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max(widest + 4, 12), 48)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_pdf(doc: ReportDocument) -> bytes:
    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), topMargin=1.4 * cm, bottomMargin=1.4 * cm,
        leftMargin=1.4 * cm, rightMargin=1.4 * cm, title=doc.title,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.textColor = _ACCENT_HEX
    subtitle_style = styles["Normal"]
    subtitle_style.textColor = colors.HexColor("#5B6474")

    elements = [
        Paragraph(doc.title, title_style),
        Paragraph(doc.tenant_name, styles["Heading3"]),
        Paragraph(doc.subtitle, subtitle_style),
        Paragraph(f"Gerado em {doc.generated_at.strftime('%d/%m/%Y %H:%M')}", subtitle_style),
        Spacer(1, 14),
    ]

    if doc.rows:
        table = Table([doc.columns, *doc.rows], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _ACCENT_HEX),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, _BORDER_HEX),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _ROW_ALT_HEX]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Nenhum registro encontrado para os filtros selecionados.", subtitle_style))

    pdf.build(elements)
    return buffer.getvalue()
